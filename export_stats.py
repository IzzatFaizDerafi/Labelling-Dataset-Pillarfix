#!/usr/bin/env python3
"""
export_stats.py
================
Reads annotations.json (COCO format) and produces:
  1. A summary table printed to the terminal
  2. An Excel report  (annotation_report.xlsx)
  3. A per-image CSV  (annotation_report.csv)

Run:
    python export_stats.py
"""

import json
import os
import sys
from pathlib import Path
from collections import defaultdict

ANNOTATIONS_FILE = r"C:\Users\Izzat\Labelling Folder\annotations.json"
OUTPUT_CSV       = r"C:\Users\Izzat\Labelling Folder\annotation_report.csv"
OUTPUT_XLSX      = r"C:\Users\Izzat\Labelling Folder\annotation_report.xlsx"


# ── Load ──────────────────────────────────────────────────────────────────────
def load_coco(path: str) -> dict:
    if not os.path.exists(path):
        sys.exit(f"[ERROR] File not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ── Build per-image rows ───────────────────────────────────────────────────────
def build_rows(coco: dict) -> list[dict]:
    cat_map = {c["id"]: c["name"] for c in coco.get("categories", [])}

    # Group annotations by image_id
    ann_by_img = defaultdict(list)
    for ann in coco.get("annotations", []):
        ann_by_img[ann["image_id"]].append(ann)

    rows = []
    for img in coco.get("images", []):
        img_id    = img["id"]
        file_name = img["file_name"]
        cls       = img.get("class", cat_map.get(
            ann_by_img[img_id][0]["category_id"] if ann_by_img[img_id] else 0, ""))
        anns      = ann_by_img[img_id]
        no_defect = img.get("no_defect", len(anns) == 0)

        sev_rank  = {"Low": 0, "Medium": 1, "High": 2}
        severities = [a.get("severity", "Low") for a in anns]
        overall   = max(severities, key=lambda s: sev_rank.get(s, 0)) \
                    if severities else ("None" if no_defect else "Low")

        # Severity percentage mapping
        sev_pct = {"None": 0, "Low": 33, "Medium": 66, "High": 100}

        low_cnt  = severities.count("Low")
        med_cnt  = severities.count("Medium")
        high_cnt = severities.count("High")

        rows.append({
            "image_id":        img_id,
            "file_name":       file_name,
            "image_name":      os.path.basename(file_name),
            "class":           cls,
            "no_defect":       no_defect,
            "num_boxes":       len(anns),
            "severity_overall": overall,
            "severity_pct":    sev_pct.get(overall, 0),
            "low_boxes":       low_cnt,
            "medium_boxes":    med_cnt,
            "high_boxes":      high_cnt,
        })
    return rows


# ── Terminal summary ──────────────────────────────────────────────────────────
def print_summary(rows: list[dict]):
    total   = len(rows)
    labeled = total   # all rows in the file are labeled

    cls_counter   = defaultdict(int)
    sev_counter   = defaultdict(int)
    box_counter   = defaultdict(int)
    no_defect_cnt = sum(1 for r in rows if r["no_defect"])

    for r in rows:
        cls_counter[r["class"]] += 1
        sev_counter[r["severity_overall"]] += 1
        box_counter[r["class"]] += r["num_boxes"]

    total_boxes = sum(r["num_boxes"] for r in rows)

    print("\n" + "═" * 55)
    print("  DEFECT DETECTION  –  ANNOTATION SUMMARY")
    print("═" * 55)
    print(f"  Total labeled images : {labeled}")
    print(f"  Images with defect   : {labeled - no_defect_cnt}")
    print(f"  No-defect images     : {no_defect_cnt}")
    print(f"  Total bounding boxes : {total_boxes}")
    print()
    print("  ── By Class ──────────────────────────────────")
    for cls, cnt in sorted(cls_counter.items()):
        boxes = box_counter[cls]
        pct   = cnt / labeled * 100 if labeled else 0
        print(f"  {cls:<25} {cnt:>5} imgs  {boxes:>5} boxes  ({pct:.1f}%)")
    print()
    print("  ── By Severity (overall per image) ───────────")
    for sev in ["None", "Low", "Medium", "High"]:
        cnt = sev_counter.get(sev, 0)
        pct = cnt / labeled * 100 if labeled else 0
        print(f"  {sev:<10} {cnt:>5} images  ({pct:.1f}%)")
    print("═" * 55 + "\n")


# ── CSV export ────────────────────────────────────────────────────────────────
def export_csv(rows: list[dict], path: str):
    import csv
    fieldnames = [
        "image_id", "image_name", "class", "no_defect",
        "num_boxes", "severity_overall", "severity_pct",
        "low_boxes", "medium_boxes", "high_boxes", "file_name",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"[CSV] Saved → {path}")


# ── Excel export ──────────────────────────────────────────────────────────────
def export_xlsx(rows: list[dict], path: str):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[XLSX] openpyxl not installed – skipping Excel export.")
        print("       Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annotations"

    # ── Header ──
    headers = [
        "Image ID", "Image Name", "Class", "No Defect",
        "# Boxes", "Severity Overall", "Severity %",
        "Low Boxes", "Medium Boxes", "High Boxes",
    ]
    header_fill = PatternFill("solid", fgColor="2D2D3F")
    header_font = Font(bold=True, color="CDD6F4")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ── Colour maps ──
    class_fills = {
        "BERKARAT":             "F5CBA7",
        "CONDONG":              "AED6F1",
        "PINTU TIDAK BERKUNCI": "D7BDE2",
        "VANDALISM":            "F1948A",
    }
    sev_fills = {
        "None":   "FFFFFF",
        "Low":    "A9DFBF",
        "Medium": "F9E79F",
        "High":   "F1948A",
    }

    for row_i, r in enumerate(rows, 2):
        vals = [
            r["image_id"], r["image_name"], r["class"],
            "Yes" if r["no_defect"] else "No",
            r["num_boxes"], r["severity_overall"], r["severity_pct"],
            r["low_boxes"], r["medium_boxes"], r["high_boxes"],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            if col == 3:   # class
                cell.fill = PatternFill("solid",
                    fgColor=class_fills.get(r["class"], "FFFFFF"))
            if col == 6:   # severity
                cell.fill = PatternFill("solid",
                    fgColor=sev_fills.get(r["severity_overall"], "FFFFFF"))

    # ── Column widths ──
    col_widths = [10, 36, 26, 11, 9, 18, 12, 11, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Summary sheet ──
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Class"
    ws2["B1"] = "Images"
    ws2["C1"] = "Bounding Boxes"
    ws2["D1"] = "% of Dataset"

    from collections import defaultdict
    cls_imgs = defaultdict(int)
    cls_boxes = defaultdict(int)
    for r in rows:
        cls_imgs[r["class"]] += 1
        cls_boxes[r["class"]] += r["num_boxes"]

    for row_i, cls in enumerate(sorted(cls_imgs), 2):
        n = cls_imgs[cls]
        ws2.cell(row=row_i, column=1, value=cls)
        ws2.cell(row=row_i, column=2, value=n)
        ws2.cell(row=row_i, column=3, value=cls_boxes[cls])
        ws2.cell(row=row_i, column=4, value=round(n / len(rows) * 100, 1))

    # Severity summary
    ws2["F1"] = "Severity"
    ws2["G1"] = "Images"
    ws2["H1"] = "% of Dataset"
    sev_imgs = defaultdict(int)
    for r in rows:
        sev_imgs[r["severity_overall"]] += 1
    for row_i, sev in enumerate(["None", "Low", "Medium", "High"], 2):
        n = sev_imgs.get(sev, 0)
        ws2.cell(row=row_i, column=6, value=sev)
        ws2.cell(row=row_i, column=7, value=n)
        ws2.cell(row=row_i, column=8, value=round(n / len(rows) * 100, 1) if rows else 0)

    wb.save(path)
    print(f"[XLSX] Saved → {path}")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    coco = load_coco(ANNOTATIONS_FILE)
    rows = build_rows(coco)

    if not rows:
        print("[INFO] No labeled images found in annotations.json yet.")
        sys.exit(0)

    print_summary(rows)
    export_csv(rows, OUTPUT_CSV)
    export_xlsx(rows, OUTPUT_XLSX)
